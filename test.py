import openpyxl
from openpyxl.drawing.image import Image

wb = openpyxl.load_workbook('./siba.xlsx')
#ws = wb.get_sheet_names()[0]
ws = wb.active
print(ws)
print(dir(ws))

for row, values in enumerate(ws.values):
    for col, value in enumerate(values):
        if value != None:
            img = Image('./.ico/siba.jpg') 
            img.height = 100
            img.width = 100
            print(ws.cell(row=row+1, column=col+1))
            p = ws.cell(row=row+1, column=col+1)
            print(dir(p),",",p.column_letter+str(p.row))
            #img.anchor(ws.cell(row=row+1, column=col+1))
            ws.add_image(img, 'A2')
            print(row+1, ", ",col+1, ", ", value)

wb.save('./siba.xlsx')

